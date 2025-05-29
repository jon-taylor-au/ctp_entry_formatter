import os
import sys
import subprocess
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Config ---
PROCESS_NAME: str = "CTP Clinical Entries Formatter"
EXCEL_FILE = Path("inputs_ctp_formatter/book_id_queue.xlsx")
STATE_FILE = Path("last_run_timestamp.txt")
LOG_DIR = Path("inputs_ctp_formatter")
LOG_FILE = LOG_DIR / "run_log.txt"
BASE_URL = os.getenv("BASE_URL")

ID_COL = "BookID"
STATUS_COL = "Status"
TIMESTAMP_COL = "Processed"
SCRIPTS = [
    "01_get_data.py",
    "02_change_data.py",
    "03_present_data.py",
    "04_write_back.py",
    "05_cleanup.py",
]


class TeeLogger:
    def __init__(self, logfile_path):
        self.terminal = sys.stdout
        self.logfile = open(logfile_path, "a", encoding="utf-8")

    def write(self, message):
        self.terminal.write(message)
        self.logfile.write(message)

    def flush(self):
        self.terminal.flush()
        self.logfile.flush()


def trim_log_file(path, max_lines=5000):
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    if len(lines) > max_lines:
        with open(path, "w", encoding="utf-8") as f:
            f.writelines(lines[-max_lines:])
        print(f"Trimmed log to last {max_lines} lines.")


def format_excel_queue(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(bold=True, color="FFFF00")
    timestamp_font = Font(bold=True, color="FFA500")
    timestamp_fill = PatternFill("solid", fgColor="FFFFFF")

    status_col = 2
    timestamp_col = 4
    columns_to_format = [1, 2, 3, 4]

    for col in columns_to_format:
        cell = ws.cell(row=1, column=col)
        if col == timestamp_col:
            cell.font = timestamp_font
            cell.fill = timestamp_fill
        else:
            cell.font = header_font
            cell.fill = header_fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = str(row[status_col - 1].value).strip().lower()
        if status == "done":
            for cell in row[:3]:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif status == "error":
            for cell in row[:3]:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")

    for col in columns_to_format:
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(file_path)
    print(f"Excel formatting applied to: {file_path}")


def run_pipeline(book_id):
    env = os.environ.copy()
    env["BOOK_ID"] = str(book_id)

    for script in SCRIPTS:
        print(f"[{book_id}] Running: {script}")
        result = subprocess.run(["python", script], env=env, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"[{book_id}] {script} failed.\n{result.stderr}")
            return "Error"
        print(f"[{book_id}] {script} completed.")
    return "Done"


def main():
    LOG_DIR.mkdir(exist_ok=True)
    sys.stdout = TeeLogger(LOG_FILE)

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print("\n" + "=" * 70)
    print(f"PROCESS: {PROCESS_NAME}")
    print(f"RUN STARTED: {now_str} in {BASE_URL}")
    print("=" * 70)

    # --- Check Excel file mtime ---
    try:
        current_mtime = EXCEL_FILE.stat().st_mtime
    except FileNotFoundError:
        print(f"Excel file not found: {EXCEL_FILE}")
        return

    if STATE_FILE.exists():
        with open(STATE_FILE, "r") as f:
            last_mtime = float(f.read().strip())
        if current_mtime == last_mtime:
            print("Excel file unchanged since last run. Exiting.")
            return

    df = pd.read_excel(EXCEL_FILE)
    df[STATUS_COL] = df[STATUS_COL].astype(str)
    df[TIMESTAMP_COL] = df[TIMESTAMP_COL].astype(str)
    print(f"Loaded {len(df)} rows from Excel.")

    changes_made = False

    for i, row in df.iterrows():
        raw_id = row.get(ID_COL)
        book_id = str(int(raw_id)).strip() if pd.notna(raw_id) else ""
        status = str(row.get(STATUS_COL)).strip().lower()

        print(f"Row {i}: BookID='{book_id}' Status='{status}'")
        if status == "done" or not book_id.isdigit():
            print("Skipping")
            continue

        print("Processing")
        result = run_pipeline(book_id)
        df.at[i, STATUS_COL] = result
        df.at[i, TIMESTAMP_COL] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{book_id}] Status updated to '{result}'")
        changes_made = True

    if changes_made:
        try:
            df.to_excel(EXCEL_FILE, index=False)
            print("Excel file updated successfully.")
            format_excel_queue(EXCEL_FILE)
        except PermissionError:
            print(f"Cannot write to {EXCEL_FILE}. Is it open?")
            return
    else:
        print("No updates made to Excel.")

    # Always update timestamp so file won't be reprocessed
    with open(STATE_FILE, "w") as f:
        f.write(str(current_mtime))
    print("Timestamp updated to reflect last checked state.")

    print("All pending Book IDs processed.")
    print(f"Run completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    trim_log_file(LOG_FILE, max_lines=5000)


if __name__ == "__main__":
    main()
